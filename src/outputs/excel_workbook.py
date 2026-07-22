"""
Excel Portfolio Workbook Generator.
Generates multi-tab Excel workbook using openpyxl, or CSV summary fallback if openpyxl is absent.
"""
import os
import csv
from typing import List, Dict, Any
from src.models.initiative import Initiative
from src.metrics.engine import MetricsEngine
from src.validation.exceptions import DataQualityException

class ExcelWorkbookGenerator:
    def generate_workbook(self, initiatives: List[Initiative], exceptions: List[DataQualityException], output_path: str) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")

            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            amber_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

            # 1. Summary Tab
            ws_sum = wb.active
            ws_sum.title = "Executive Summary"
            ws_sum["A1"] = "ENTERPRISE PORTFOLIO EXECUTIVE SUMMARY"
            ws_sum["A1"].font = title_font
            ws_sum["A2"] = "SYNTHETIC DEMONSTRATION DATA — NOT FROM A REAL ORGANISATION"
            ws_sum["A2"].font = Font(italic=True, color="7F7F7F")

            engine = MetricsEngine()
            m = engine.calculate_portfolio_summary(initiatives)

            headers_sum = ["Metric", "Value"]
            ws_sum.append([])
            ws_sum.append(headers_sum)
            for col_num in range(1, 3):
                cell = ws_sum.cell(row=4, column=col_num)
                cell.font = header_font
                cell.fill = header_fill

            metrics_data = [
                ("Total Initiatives", m["total_initiatives"]),
                ("Overall RAG Status", m["overall_rag"]),
                ("Portfolio Health Score", f"{m['health_score']}/100"),
                ("Total Approved Budget (£)", f"£{m['total_budget']:,.2f}"),
                ("Total Actual Spend (£)", f"£{m['total_actual_cost']:,.2f}"),
                ("Spend Variance (%)", f"{m['spend_variance_pct']}%"),
                ("RED Initiatives Count", m["red_count"]),
                ("AMBER Initiatives Count", m["amber_count"]),
                ("GREEN Initiatives Count", m["green_count"]),
            ]

            for r_idx, (metric, val) in enumerate(metrics_data, start=5):
                ws_sum.cell(row=r_idx, column=1, value=metric)
                c2 = ws_sum.cell(row=r_idx, column=2, value=val)
                if val == "RED": c2.fill = red_fill
                elif val == "AMBER": c2.fill = amber_fill
                elif val == "GREEN": c2.fill = green_fill

            # 2. Initiatives Tab
            ws_init = wb.create_sheet(title="Initiative Register")
            init_headers = [
                "Initiative ID", "Title", "Category", "Stage", "Priority",
                "RAG Status", "Progress %", "Budget (£)", "Actual Spend (£)",
                "Forecast (£)", "Delivery Owner", "Business Owner", "Data Quality"
            ]
            ws_init.append(init_headers)
            for col_num in range(1, len(init_headers) + 1):
                cell = ws_init.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill

            for init in initiatives:
                row_val = [
                    init.initiative_id, init.title, init.portfolio_category,
                    init.lifecycle_stage, init.priority, init.rag_status,
                    init.progress_pct, init.budget, init.actual_cost,
                    init.forecast_cost, init.delivery_owner, init.business_owner,
                    init.data_quality_status
                ]
                ws_init.append(row_val)
                r_idx = ws_init.max_row
                rag_cell = ws_init.cell(row=r_idx, column=6)
                if init.rag_status == "RED": rag_cell.fill = red_fill
                elif init.rag_status == "AMBER": rag_cell.fill = amber_fill
                elif init.rag_status == "GREEN": rag_cell.fill = green_fill

            # 3. Data Quality Exceptions Tab
            ws_dq = wb.create_sheet(title="DQ Exceptions")
            dq_headers = ["Rule ID", "Rule Name", "Severity", "Initiative ID", "Field", "Error Message", "Owner"]
            ws_dq.append(dq_headers)
            for col_num in range(1, len(dq_headers) + 1):
                cell = ws_dq.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill

            for exc in exceptions:
                ws_dq.append([
                    exc.rule_id, exc.rule_name, exc.severity,
                    exc.initiative_id, exc.field_name, exc.error_message, exc.owner
                ])

            # Auto-fit column widths across all sheets
            for sheet in wb.worksheets:
                sheet.freeze_panes = "A2"
                for col in sheet.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            return output_path
        except ImportError:
            # Fallback CSV format if openpyxl is not installed
            csv_path = output_path.replace(".xlsx", ".csv")
            os.makedirs(os.path.dirname(csv_path), exist_ok=True)
            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Initiative ID", "Title", "Category", "Stage", "RAG Status", "Budget", "Actual Spend"])
                for init in initiatives:
                    writer.writerow([init.initiative_id, init.title, init.portfolio_category, init.lifecycle_stage, init.rag_status, init.budget, init.actual_cost])
            return csv_path
